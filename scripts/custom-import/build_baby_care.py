"""
Convert 俄语婴幼儿护理词汇表.xlsx → JSON seed file for the Earthworm course pack system.

Output: backend/src/main/resources/customs/baby_care.json
Schema:
{
  "id": "ru-baby-care",
  "title": "...",
  "description": "...",
  "cover": "https://...",
  "courses": [
    { "id": "...", "title": "...", "description": "...",
      "statements": [{ "english": "...", "chinese": "...", "soundmark": "..." }] }
  ]
}

Each sheet → multiple courses of LESSON_SIZE (=10) statements.
Each row becomes a statement:
  english   = Russian word (col 0)
  chinese   = Chinese gloss (col 3)
  soundmark = "<phonetic hint> ｜ <example sentence>"

Usage:
    pip install openpyxl
    python build_baby_care.py [xlsx_path]
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from pathlib import Path

import openpyxl

LESSON_SIZE = 10
PACK_ID = "ru-baby-care"
PACK_TITLE = "婴幼儿护理俄语 · 实用词汇与例句"
PACK_DESCRIPTION = (
    "面向中文母语父母 / 育儿嫂 / 月嫂学习俄语的实用词汇与例句集。"
    "覆盖身体部位、日常护理、辅食喂养、健康安全、紧急情况、沟通用语等 11 大场景。"
)
PACK_COVER = "https://images.unsplash.com/photo-1519689680058-324335c77eba?q=80&w=1200&auto=format&fit=crop"
DEFAULT_INPUT = Path(r"C:\Users\Administrator\Desktop\俄语婴幼儿护理词汇表.xlsx")
DEFAULT_OUTPUT = Path(__file__).resolve().parents[2] / "backend/src/main/resources/customs/baby_care.json"


def short_hash(s: str) -> str:
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:8]


def normalize_sheet_name(s: str) -> str:
    return re.sub(r"\s+", "", s.strip())


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", nargs="?", default=str(DEFAULT_INPUT))
    ap.add_argument("--output", default=str(DEFAULT_OUTPUT))
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.output)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    print(f">>> loading {xlsx_path} ...", file=sys.stderr)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    courses: list[dict] = []
    course_order = 0
    total_rows = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header = [str(c or "").strip() for c in rows[0]]
        try:
            ru_idx = header.index("俄语单词")
            zh_idx = header.index("中文释义")
        except ValueError:
            print(f"  skip {sheet_name}: missing required columns", file=sys.stderr)
            continue
        # Optional columns
        ipa_idx = header.index("音标/发音提示") if "音标/发音提示" in header else None
        ex_idx = header.index("例句") if "例句" in header else None

        items: list[dict] = []
        for r in rows[1:]:
            if not r:
                continue
            ru = (str(r[ru_idx]) if r[ru_idx] is not None else "").strip()
            zh = (str(r[zh_idx]) if r[zh_idx] is not None else "").strip()
            if not ru or not zh:
                continue
            ipa = (str(r[ipa_idx]) if ipa_idx is not None and r[ipa_idx] is not None else "").strip()
            ex = (str(r[ex_idx]) if ex_idx is not None and r[ex_idx] is not None else "").strip()
            sm_parts = [p for p in [ipa, ex] if p]
            soundmark = " ｜ ".join(sm_parts) or ru
            items.append({
                "english": ru,
                "chinese": zh,
                "soundmark": soundmark,
            })

        total_rows += len(items)
        # Split into lessons of LESSON_SIZE
        s_norm = normalize_sheet_name(sheet_name)
        lesson_idx = 0
        for i in range(0, len(items), LESSON_SIZE):
            lesson_idx += 1
            course_order += 1
            chunk = items[i:i + LESSON_SIZE]
            lo = i + 1
            hi = i + len(chunk)
            course_id = f"{PACK_ID}-{s_norm}-{lesson_idx}-{short_hash(sheet_name + str(lesson_idx))}"
            statements = []
            for k, it in enumerate(chunk, start=1):
                statements.append({
                    "english": it["english"],
                    "chinese": it["chinese"],
                    "soundmark": it["soundmark"],
                })
            courses.append({
                "id": course_id,
                "order": course_order,
                "title": f"{sheet_name} · 第 {lesson_idx} 课（{lo}-{hi}）",
                "description": f"{sheet_name} 主题词汇 {lo}-{hi}，看中文输入俄语单词；soundmark 含发音提示和原例句。",
                "statements": statements,
            })
        print(f"  {sheet_name}: {len(items)} items → {lesson_idx} courses", file=sys.stderr)

    pack = {
        "id": PACK_ID,
        "title": PACK_TITLE,
        "description": f"{PACK_DESCRIPTION}\n\n共 {total_rows} 词，{len(courses)} 节课。",
        "cover": PACK_COVER,
        "isFree": True,
        "courses": courses,
    }

    out_path.write_text(json.dumps(pack, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f">>> wrote {out_path} ({len(courses)} courses, {total_rows} statements)")


if __name__ == "__main__":
    main()
