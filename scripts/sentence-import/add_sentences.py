"""
Match Tatoeba RU-ZH pairs to all target word lists and append
"sentence courses" to each target JSON file.

Target files:
  - backend/src/main/resources/torfl/levels/{A1..C2}.json  (in-place "sentences" extension)
  - backend/src/main/resources/customs/baby_care.json      (extract embedded examples into new courses)
  - backend/src/main/resources/customs/oil_engineering.json (no Tatoeba match expected for technical terms)

Logic:
  1. Lemmatize every Tatoeba RU sentence (pymorphy3) → build lemma → sentence_id index.
  2. For each lemma in target word lists, pick up to N (=2) sentences that:
        - contain the lemma
        - 4-14 tokens long (sweet spot for learners)
        - have a non-trivial ZH translation
  3. For TORFL levels: append sentences into level JSON's "sentences" array.
  4. For baby_care: existing example sentences (already in row data) are gathered
     into "sentence courses" prepended; Tatoeba sentences are a secondary supplement.

Run:
    pip install pymorphy3 pymorphy3-dicts-ru
    python add_sentences.py
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pymorphy3

HERE = Path(__file__).parent
ROOT = HERE.parent.parent
LEVELS_DIR = ROOT / "backend/src/main/resources/torfl/levels"
CUSTOMS_DIR = ROOT / "backend/src/main/resources/customs"
TATOEBA_PATH = HERE / "output/ru_zh_pairs.jsonl"
BABY_CARE_XLSX = Path(r"C:\Users\Administrator\Desktop\俄语婴幼儿护理词汇表.xlsx")

PER_WORD_LIMIT = 2  # at most N sentences per word
MIN_TOKENS, MAX_TOKENS = 4, 14
SENTENCES_PER_LEVEL_CAP = 200  # don't blow up A1 with thousands

CYRILLIC_TOKEN = re.compile(r"[а-яёА-ЯЁ][а-яёА-ЯЁ\-]*")


def load_tatoeba():
    pairs = []
    with TATOEBA_PATH.open("r", encoding="utf-8") as f:
        for line in f:
            try:
                pairs.append(json.loads(line))
            except Exception:
                continue
    print(f">>> loaded {len(pairs)} RU-ZH pairs", file=sys.stderr)
    return pairs


def lemma_index(pairs, morph: pymorphy3.MorphAnalyzer):
    """Build lemma → list of pair indices."""
    idx = defaultdict(list)
    for i, p in enumerate(pairs):
        tokens = CYRILLIC_TOKEN.findall(p["ru"])
        if not (MIN_TOKENS <= len(tokens) <= MAX_TOKENS):
            continue
        # Skip overly long ZH (likely paragraph)
        if len(p["zh"]) > 60:
            continue
        seen_lemmas = set()
        for t in tokens:
            lemma = morph.parse(t)[0].normal_form
            if lemma and lemma not in seen_lemmas:
                seen_lemmas.add(lemma)
                idx[lemma].append(i)
        if (i + 1) % 2000 == 0:
            print(f"    indexed {i+1}/{len(pairs)} pairs ({len(idx)} unique lemmas)", file=sys.stderr)
    print(f">>> lemma index built: {len(idx)} unique lemmas", file=sys.stderr)
    return idx


def pick_sentences(lemma: str, pairs, idx, used_ids: set, limit: int):
    """Pick up to `limit` sentences for a lemma, preferring short & unseen."""
    candidates = idx.get(lemma, [])
    if not candidates:
        return []
    scored = []
    for pid in candidates:
        if pid in used_ids:
            continue
        p = pairs[pid]
        ru_tokens = len(CYRILLIC_TOKEN.findall(p["ru"]))
        # Score: prefer 6-10 tokens, short ZH, no rare punctuation
        score = abs(8 - ru_tokens) * 2 + (max(0, len(p["zh"]) - 25))
        scored.append((score, pid, p))
    scored.sort(key=lambda x: x[0])
    out = []
    for _, pid, p in scored[:limit]:
        used_ids.add(pid)
        out.append({"russian": p["ru"], "chinese": p["zh"]})
    return out


def update_torfl_levels(pairs, idx):
    """Extend each level's `sentences` array with Tatoeba matches for that level's words."""
    levels = ["A1", "A2", "B1", "B2", "C1", "C2"]
    used = set()  # don't repeat sentences across levels
    summary = {}
    for lvl in levels:
        path = LEVELS_DIR / f"{lvl}.json"
        if not path.exists():
            continue
        obj = json.loads(path.read_text(encoding="utf-8"))
        # Words at this level — keep order
        words = obj.get("words", [])
        existing_sentences = obj.get("sentences", [])
        existing_keys = {s.get("russian", "").strip() for s in existing_sentences}
        added = 0
        new_sentences = list(existing_sentences)
        for w in words:
            if added >= SENTENCES_PER_LEVEL_CAP:
                break
            lemma = (w.get("russian") or "").lower()
            if not lemma:
                continue
            for s in pick_sentences(lemma, pairs, idx, used, PER_WORD_LIMIT):
                if s["russian"] in existing_keys:
                    continue
                existing_keys.add(s["russian"])
                new_sentences.append(s)
                added += 1
                if added >= SENTENCES_PER_LEVEL_CAP:
                    break
        obj["sentences"] = new_sentences
        # Update description
        desc = obj.get("description", "")
        desc = re.sub(r"\s*\+\s*\d+\s*例句来自 Tatoeba.*$", "", desc).strip()
        desc = desc + f"\n+ {added} 例句来自 Tatoeba 俄汉句对（母语人编写，语法保证）。"
        obj["description"] = desc
        path.write_text(json.dumps(obj, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        summary[lvl] = {"existing": len(existing_sentences), "added": added, "total": len(new_sentences)}
        print(f"  {lvl}: +{added} sentences (total {len(new_sentences)})", file=sys.stderr)
    return summary


# ---- baby_care: extract built-in examples into dedicated sentence courses ----

def slug(s: str) -> str:
    return re.sub(r"[^a-zA-Z0-9_\-]+", "", s)


def short_hash(s: str) -> str:
    import hashlib
    return hashlib.sha1(s.encode("utf-8")).hexdigest()[:8]


def baby_care_sentence_courses():
    """Read xlsx directly to get original Russian sentence + Chinese translation per row.
    Group by sheet, split into lessons of 10 sentences each. Append to baby_care.json."""
    path = CUSTOMS_DIR / "baby_care.json"
    pack = json.loads(path.read_text(encoding="utf-8"))
    existing_ids = {c.get("id") for c in pack.get("courses", [])}
    if not BABY_CARE_XLSX.exists():
        print(f"  skip baby_care sentence extract: {BABY_CARE_XLSX} missing", file=sys.stderr)
        return {"added_courses": 0, "added_sentences": 0}

    wb = openpyxl.load_workbook(BABY_CARE_XLSX, read_only=True, data_only=True)
    pack_id = pack["id"]
    next_order = max((c.get("order", 0) for c in pack["courses"]), default=0)
    added_courses = 0
    added_sentences = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue
        header = [str(c or "").strip() for c in rows[0]]
        if "例句" not in header:
            continue
        ex_idx = header.index("例句")
        zh_idx = header.index("中文释义") if "中文释义" in header else None

        items = []
        for r in rows[1:]:
            if not r or len(r) <= ex_idx:
                continue
            ex = (str(r[ex_idx]) if r[ex_idx] is not None else "").strip()
            if not ex:
                continue
            # Examples are formatted like "Умываем голову. (洗头。)"
            m = re.match(r"^(.+?)\s*[\(（](.+?)[\)）]\s*$", ex)
            if m:
                ru = m.group(1).strip()
                zh = m.group(2).strip()
            else:
                ru = ex
                zh = (str(r[zh_idx]) if zh_idx is not None and r[zh_idx] is not None else "").strip()
            if not ru or not zh:
                continue
            items.append({"english": ru, "chinese": zh, "soundmark": ""})

        if not items:
            continue
        s_slug = slug(sheet_name) or short_hash(sheet_name)
        for i in range(0, len(items), 10):
            chunk = items[i:i + 10]
            next_order += 1
            lesson_num = i // 10 + 1
            course_id = f"{pack_id}-{s_slug}-sent-{lesson_num}-{short_hash(sheet_name + '-sent-' + str(lesson_num))}"
            if course_id in existing_ids:
                continue
            pack["courses"].append({
                "id": course_id,
                "order": next_order,
                "title": f"{sheet_name} · 例句 第 {lesson_num} 课（{i+1}-{i+len(chunk)}）",
                "description": f"{sheet_name} 主题例句 {i+1}-{i+len(chunk)}：看中文输入完整俄语句子（母语原例）。",
                "statements": chunk,
            })
            existing_ids.add(course_id)
            added_courses += 1
            added_sentences += len(chunk)

    # Update pack description
    pack["description"] = re.sub(r"\s*\+\s*\d+ 例句课.*$", "", pack["description"]).strip()
    pack["description"] += f"\n+ {added_courses} 节例句课 / {added_sentences} 条句子（提取自原 xlsx 的母语例句）。"
    path.write_text(json.dumps(pack, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"  baby_care: +{added_courses} sentence courses / {added_sentences} sentences", file=sys.stderr)
    return {"added_courses": added_courses, "added_sentences": added_sentences}


def main() -> None:
    morph = pymorphy3.MorphAnalyzer()
    pairs = load_tatoeba()
    idx = lemma_index(pairs, morph)

    summary = {}
    summary["torfl"] = update_torfl_levels(pairs, idx)
    summary["baby_care"] = baby_care_sentence_courses()

    print("\nSummary:")
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
